#List of installs/libraries needed (ensure you have installed all of them)
import psycopg2
import pandas as pd
import sys
import xlsxwriter
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def connect(): #Connect to postgres database, this function will ensure you are successfully authenticated and logged in
    conn = None
    #Below we are testing if you have all the correct data (This is a data privacy measure, so we want to make sure it is successful)
    try:
        print('Connecting...')
        conn = psycopg2.connect(database="SEI DB", #change this to whatever you named your database!
                                host="localhost",
                                user="postgres",
                                password="password", #insert whatever password you created for your server
                                port="5432") #you can check the port number by right clicking on the server in Postgres and looking at the properties. However this shouldn't change and it should be 5432.
    except(Exception, psycopg2.DatabaseError) as error: 
        print(error)
        sys.exit(1) #If you have an error, there is an issue connecting to your Postgres database. Make sure you are logged into your server in pgAdmin
    print('All good, Connection successful!')

    return conn


def SQLtoDF(conn): #Import data from postgres, this function is what is allowing you to import all your data. We need to make sure you are following the database structure
    cursor = conn.cursor()
    query = cursor.execute("SELECT * FROM public.\"tableName\"") #switch "tableName" to whatever your table is named in postgres
    try: 
        cursor.execute(query)
    except (Exception, psycopg2.DatabaseError) as error:
        print("Error: %s" % error)
    
    #All column names from linkedin sales nav csv files, this shouldnt change unless you change the data structure of the spreadsheet you imported
    columnNames = ["profileUrl", "fullName", "firstName", "lastName", "companyName", "title", "companyID", "companyUrl", "regularCompanyUrl", "summary", "titleDescription", "industry", "companyLocation", "location", "durationInRole", "durationInCompany", "connectionDegree", "profileImage", "sharedConnectionsCount", "name", "vmid", "linkedInProfileUrl", "isPremium", "isOpenLink", "query", "timestamp", "duration"]

    #The execute returns a list of tuples
    tuplesList = cursor.fetchall()
    cursor.close()

    df = pd.DataFrame(tuplesList, columns = columnNames)
    return df

def allContacts(df): #This function is processing the data and getting rid of the information you dont really need. If you want to include any other columns from the database, this is where you would alter the code to do that
    selected_columns = ["profileUrl", "fullName", "sharedConnectionsCount", "companyName", "title", "companyLocation", "location", "linkedInProfileUrl"]
    data = df[selected_columns].values.tolist()

    return data


##IMPORTANT: lines 55-70 and 134-149 are commented out because it is what I used to find top companies. You can uncomment by simply deleteing the pound symbols (#) in front of each line.
#It is what I used for my second use-case in the presentation!


#def get_top_companies(df, n=10, excludeCompany=None): #This function is filtering through the companies from the database and determining the strongest network presence of each company
    #if excludeCompany:
    #    df_filtered = df[df["companyName"] != excludeCompany]
    #else:
    #    df_filtered = df

    #company_counts = df_filtered.groupby("companyName")["sharedConnectionsCount"].max().sort_values(ascending=False).head(n)

    #return company_counts

#def get_company_contacts(df, top_companies): #This function is processing the data and getting rid of useless information. Again this is the function you would tweak if you want any new information in your output
    #selected_columns = ["fullName", "sharedConnectionsCount", "companyName", "title", "companyLocation", "location", "linkedInProfileUrl"]
    
    #contacts = df[df["companyName"].isin(top_companies.index)][selected_columns].values.tolist()    

    #return contacts

def save_to_excel(contacts, top_companies, output_file): #This function is all about the structure of the output excel file. THIS FUNCTION IS ALSO FOR THE SECOND COMPANY CONTACTS USE-CASE. (NOT USED WHEN LINES 55-70 and 134-149 ARE COMMENTED OUT)
    df = pd.DataFrame(contacts, columns = ["profileUrl", "fullName", "sharedConnectionsCount", "companyName", "title", "companyLocation", "location", "linkedInProfileUrl"])
    df_sorted = df.sort_values(by="companyName") #You can change this to however you want to sort the output

    # Create an Excel writer using openpyxl as the engine
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    df_sorted.to_excel(writer, index=False, sheet_name='Contacts')

    # Access the workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Contacts']

    # Apply conditional formatting based on shared connections count
    num_contacts = len(contacts)
    fill_red = PatternFill(start_color="f26e49", end_color="f26e49", fill_type="solid")
    fill_yellow = PatternFill(start_color="eff249", end_color="eff249", fill_type="solid")
    fill_green = PatternFill(start_color="12a102", end_color="12a102", fill_type="solid")

    for row in range(2, num_contacts + 2):
        shared_connections = worksheet[f'B{row}'].value

        cell = worksheet[f'A{row}']
        if shared_connections <= 2:
            cell.fill = fill_red
        elif 3 <= shared_connections <= 5:
            cell.fill = fill_yellow
        else:
            cell.fill = fill_green

    # Add the top companies and shared connections count to the Excel sheet
    top_companies.to_excel(writer, sheet_name='Contacts', startrow=num_contacts + 4, index=True, header=True)

    # Save the Excel file
    writer._save()

def networkAnalysis(data, networkAnalyisFile): #This is our network analysis function. 
    
    df = pd.DataFrame({"fullName": [contact[1] for contact in data],
                       "sharedConnectionsCount": [contact[2] for contact in data],
                       "companyName": [contact[3] for contact in data],
                       "title": [contact[4] for contact in data],
                       "companyLocation": [contact[5] for contact in data],
                       "location": [contact[5] for contact in data],
                       "linkedInProfileUrl": [contact[7] for contact in data],
                       "profileUrl": [contact[0] for contact in data],
                       })

    df_sorted = df.sort_values(by="sharedConnectionsCount", ascending=False) #We can change this line to however you want the code sorted!

    writer = pd.ExcelWriter(networkAnalyisFile, engine='xlsxwriter')
    df_sorted.to_excel(writer, index=False, sheet_name='networkAnalysis')


    writer._save()
    print(f"Network Analysis calculated and saved to {networkAnalyisFile}")


def main(): #This is the output (what you see when you run it)
    conn = connect()
    df = SQLtoDF(conn)
    data = allContacts(df)

    #print("---------------------------------------------------------------------------")

    #top_companies = get_top_companies(df, excludeCompany="Lowe\'s Companies, Inc." )
    #print("Top Companies:")
    #for i in range(len(top_companies)):
    #    company = top_companies.index[i]
    #    shared_connections = top_companies.values[i]
    #    print(f"Company: {company}, Shared Connections: {shared_connections}")

    #company_contacts = get_company_contacts(df, top_companies)
    
    print("---------------------------------------------------------------------------")

    #output_file = "company_contacts.xlsx"
    #save_to_excel(company_contacts, top_companies, output_file)
    #print(f"Company contacts saved to {output_file}")

    networkAnalyisFile = "networkAnalysis.xlsx"
    networkAnalysis(data, networkAnalyisFile)

    print("---------------------------------------------------------------------------")


    


#Don't change this
if __name__ == "__main__":
    main()








